"""Pure parsing helpers for RBA CSV / XLSX / legacy-XLS payloads (no network).

Two symmetric jobs:

* **Metadata** (the enumerator/catalog feed) — extract one descriptive row per
  ``Series ID`` from the fixed-shape header block at the top of every CSV file or
  workbook sheet (``Title`` / ``Description`` / ``Frequency`` / ``Units`` /
  ``Series ID``).
* **Data** (the fetch path) — melt a CSV file *or* a workbook sheet's data rows
  (date in column 0, one value column per series) into the long ``[table_id, title,
  date, value, series_key]`` frame. ``_melt_sheet_rows`` is the row-matrix twin of
  ``_parse_rba_csv`` so XLSX (openpyxl) and legacy XLS (xlrd) sheets fetch the same
  way a CSV does — this is what makes the XLSX-exclusive + xls-hist catalog rows
  fetchable, closing the prior "catalog ⊋ connector" gap.
"""

from __future__ import annotations

import csv
import io
from datetime import date as _date
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd
import xlrd

_CATEGORY_PREFIXES = {
    "a": "Reserve Bank",
    "b": "Banking and Finance",
    "c": "Credit and Charge Cards",
    "d": "Monetary Aggregates",
    "e": "Household and Business Finance",
    "f": "Interest Rates and Yields",
    "g": "Exchange Rates",
    "h": "Economic Activity",
    "i": "Balance of Payments",
}

_FETCH_COLUMNS = ["table_id", "title", "date", "value", "series_key"]


def _is_nondata_sheet(name: str) -> bool:
    """True for RBA's known non-data sheets (Notes / Series breaks / etc.).

    Mirrors readrba's exclusion set, matched case-insensitively on substrings so
    ``"Notes "``, ``"AGS - Notes"`` and ``"Use of Expert Judgement"`` are all caught.
    """
    low = name.lower()
    return "note" in low or "series breaks" in low or "expert judgement" in low or "expert judgment" in low


# ---------------------------------------------------------------------------
# Date / value coercion
# ---------------------------------------------------------------------------


def _normalize_date(s: str) -> str:
    for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _normalize_date_cell(cell: Any) -> str:
    """Normalize a period cell that may be a datetime object (xlsx/xls) or a string."""
    if isinstance(cell, (datetime, _date)):
        return cell.strftime("%Y-%m-%d")
    return _normalize_date(str(cell).strip())


def _to_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# CSV — data fetch
# ---------------------------------------------------------------------------


def _parse_rba_csv(text: str, table_id: str) -> pd.DataFrame:
    """Parse an RBA CSV: skip the metadata header, melt to long format."""
    lines = text.strip().split("\n")

    header_idx = 0
    series_id_row: dict[str, str] = {}

    for i, line in enumerate(lines):
        lower = line.strip().lower()
        if lower.startswith("series id"):
            reader = csv.reader(io.StringIO(line))
            parts = list(reader)[0]
            for j, part in enumerate(parts):
                series_id_row[str(j)] = part.strip()
        if lower.startswith("title"):
            header_idx = i

    if header_idx == 0:
        header_idx = min(10, len(lines) - 2)

    data_text = "\n".join(lines[header_idx:])
    rows_list = list(csv.reader(io.StringIO(data_text)))

    if len(rows_list) < 2:
        return pd.DataFrame(columns=_FETCH_COLUMNS)

    header = rows_list[0]
    data_start = 1
    for i, row in enumerate(rows_list):
        if row and row[0].strip().lower().startswith("series id"):
            data_start = i + 1
            break

    all_rows: list[dict[str, Any]] = []

    for row in rows_list[data_start:]:
        if not row or not row[0].strip():
            continue
        date = _normalize_date(row[0].strip())
        for col_idx in range(1, min(len(header), len(row))):
            col_name = header[col_idx].strip()
            if not col_name:
                continue
            all_rows.append(
                {
                    "table_id": table_id,
                    "title": col_name,
                    "date": date,
                    "value": _to_float(row[col_idx].strip()),
                    "series_key": series_id_row.get(str(col_idx), col_name),
                }
            )

    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame(columns=_FETCH_COLUMNS)


# ---------------------------------------------------------------------------
# Workbook sheet — data fetch (the row-matrix twin of _parse_rba_csv)
# ---------------------------------------------------------------------------


def _melt_sheet_rows(rows: list[list[Any]], table_id: str) -> pd.DataFrame:
    """Melt one workbook sheet (already read into a row matrix) to long format.

    ``rows`` is every row of a single data sheet — a metadata header block
    (``Title`` ... ``Series ID``/``Mnemonic`` at column 0) followed by data rows
    (a period in column 0, one value per series column). Used by ``rba_fetch`` for
    XLSX-exclusive sheets (openpyxl rows) and legacy xls-hist sheets (xlrd rows).
    """
    title_row: list[Any] | None = None
    id_row: list[Any] | None = None
    id_row_idx: int | None = None
    for i, row in enumerate(rows):
        if not row:
            continue
        head = str(row[0]).strip() if row[0] is not None else ""
        if head == "Title":
            title_row = row
        elif head in ("Series ID", "Mnemonic"):
            id_row = row
            id_row_idx = i

    if id_row is None or id_row_idx is None:
        return pd.DataFrame(columns=_FETCH_COLUMNS)

    def _cell(row: list[Any] | None, col: int) -> str:
        if row is None or col >= len(row) or row[col] is None:
            return ""
        s = str(row[col]).strip()
        return "" if s.lower() == "nan" else s

    n_cols = len(id_row)
    all_rows: list[dict[str, Any]] = []
    for row in rows[id_row_idx + 1 :]:
        if not row or row[0] is None or not str(row[0]).strip():
            continue
        date = _normalize_date_cell(row[0])
        for col in range(1, min(n_cols, len(row))):
            sid = _cell(id_row, col)
            if not sid:
                continue
            title = _cell(title_row, col) or sid
            all_rows.append(
                {
                    "table_id": table_id,
                    "title": title,
                    "date": date,
                    "value": _to_float(row[col]),
                    "series_key": sid,
                }
            )

    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame(columns=_FETCH_COLUMNS)


def _xlsx_full_sheet_rows(data: bytes, sheet_name: str) -> list[list[Any]]:
    """Every row of an XLSX sheet as plain lists (openpyxl, values only)."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _xls_full_sheet_rows(data: bytes, sheet_name: str) -> list[list[Any]]:
    """Every row of a legacy XLS sheet as plain lists (xlrd), converting Excel date
    serials in any cell to ``datetime`` so the period column parses correctly."""
    wb = xlrd.open_workbook(file_contents=data)
    if sheet_name not in wb.sheet_names():
        return []
    sheet = wb.sheet_by_name(sheet_name)
    out: list[list[Any]] = []
    for r in range(sheet.nrows):
        cells: list[Any] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    cells.append(xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode))
                except Exception:
                    cells.append(cell.value)
            else:
                cells.append(cell.value)
        out.append(cells)
    return out


def _xls_data_sheet_names(data: bytes) -> list[str]:
    """Data sheet names of a legacy XLS workbook (excluding Notes / Series breaks)."""
    try:
        wb = xlrd.open_workbook(file_contents=data, on_demand=True)
    except Exception:
        return []
    return [sn for sn in wb.sheet_names() if not _is_nondata_sheet(sn)]


# ---------------------------------------------------------------------------
# Metadata — shared header-row extraction (enumerator/catalog feed)
# ---------------------------------------------------------------------------


def _metadata_from_header_rows(
    sheet_rows: list[list[Any]],
    *,
    table_id: str,
    sheet_name: str,
    source: str,
    category: str,
) -> list[dict[str, str]]:
    """Extract series metadata from RBA-shaped header rows.

    Shared post-parsing step for both XLSX (openpyxl) and XLS-hist (xlrd) workbooks.
    ``sheet_rows`` is the first ~20 rows of a sheet, each a list of cell values with
    column 0 carrying the label (``Title``, ``Description``, ..., ``Series ID`` /
    ``Mnemonic``). One catalog row per non-empty id column.
    """
    title_row: list[Any] | None = None
    description_row: list[Any] | None = None
    frequency_row: list[Any] | None = None
    units_row: list[Any] | None = None
    id_row: list[Any] | None = None
    for row in sheet_rows[:20]:
        if not row:
            continue
        head = str(row[0]).strip() if row[0] is not None else ""
        if head == "Title":
            title_row = row
        elif head == "Description":
            description_row = row
        elif head == "Frequency":
            frequency_row = row
        elif head == "Units":
            units_row = row
        elif head in ("Series ID", "Mnemonic"):
            id_row = row

    if not id_row:
        return []

    def _cell(row: list[Any] | None, col: int) -> str:
        if row is None or col >= len(row):
            return ""
        val = row[col]
        if val is None:
            return ""
        s = str(val).strip()
        return "" if s.lower() == "nan" else s

    rows: list[dict[str, str]] = []
    for col in range(1, len(id_row)):
        sid = _cell(id_row, col)
        if not sid:
            continue
        title = _cell(title_row, col) or sid
        description = _cell(description_row, col)
        frequency = _cell(frequency_row, col)
        unit = _cell(units_row, col)
        # When a workbook has multiple data sheets, disambiguate by folding the sheet
        # name into the table_id so rows stay unique even if the same series id is
        # duplicated across sheets (and so rba_fetch can re-open the right sheet).
        effective_table_id = table_id if not sheet_name else f"{table_id}/{sheet_name}"
        rows.append(
            {
                "code": f"{effective_table_id}#{sid}",
                "table_id": effective_table_id,
                "series_id": sid,
                "title": title,
                "description": description,
                "category": category,
                "frequency": frequency,
                "unit": unit,
                "source": source,
            }
        )
    return rows


def _xlsx_sheet_rows(wb: openpyxl.Workbook, sheet_name: str) -> list[list[Any]]:
    """First 20 rows of an XLSX sheet as plain lists of cell values."""
    ws = wb[sheet_name]
    rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 20:
            break
        rows.append(list(row))
    return rows


def _xls_sheet_rows(sheet: xlrd.sheet.Sheet) -> list[list[Any]]:
    """First 20 rows of a legacy .xls sheet as plain lists."""
    rows: list[list[Any]] = []
    for r in range(min(20, sheet.nrows)):
        rows.append(list(sheet.row_values(r)))
    return rows


def _parse_xlsx_workbook_exclusive(
    data: bytes,
    stem: str,
    covered_ids: set[str],
) -> list[dict[str, str]]:
    """Emit metadata rows for a current XLSX workbook's series **not already covered**
    by the CSV pass (dynamic exclusivity — replaces a hardcoded sheet allow-list).

    Walks every data sheet; a series whose id is already present in ``covered_ids``
    (the union of CSV-derived series ids) is dropped. In practice this yields exactly
    ``a03.xlsx`` → "Bond Purchase Program" today, but it self-maintains: any future
    XLSX-only sheet is picked up automatically. Tagged ``source='rba_xlsx'``.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return []
    category = _CATEGORY_PREFIXES.get(stem[0].lower() if stem else "", "")
    rows: list[dict[str, str]] = []
    try:
        for sheet_name in wb.sheetnames:
            if _is_nondata_sheet(sheet_name):
                continue
            sheet_rows = _xlsx_sheet_rows(wb, sheet_name)
            for row in _metadata_from_header_rows(
                sheet_rows,
                table_id=stem,
                sheet_name=sheet_name,
                source="rba_xlsx",
                category=category,
            ):
                if row["series_id"] not in covered_ids:
                    rows.append(row)
    finally:
        wb.close()
    return rows


def _parse_xls_hist(data: bytes, table_id: str) -> list[dict[str, str]]:
    """Parse a legacy ``.xls`` workbook from ``/statistics/tables/xls-hist/``.

    Walks every data sheet (skipping Notes / Series breaks) and emits one row per
    series id column. Series are suffixed with the sheet name to keep keys unique when
    a workbook has multiple data sheets (e.g. ``a03hist-2003-2008.xls`` has one sheet
    per security). The legacy ``zcr-analytical-series-hist`` workbook labels its series
    ``Mnemonic`` instead of ``Series ID`` — handled by the shared header parser.
    """
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception:
        return []
    category = _CATEGORY_PREFIXES.get(table_id[0].lower() if table_id else "", "")

    rows: list[dict[str, str]] = []
    data_sheets = [sn for sn in wb.sheet_names() if not _is_nondata_sheet(sn)]
    attach_sheet = len(data_sheets) > 1
    for sheet_name in data_sheets:
        sheet = wb.sheet_by_name(sheet_name)
        sheet_rows = _xls_sheet_rows(sheet)
        rows.extend(
            _metadata_from_header_rows(
                sheet_rows,
                table_id=table_id,
                sheet_name=sheet_name if attach_sheet else "",
                source="rba_xlsx_hist",
                category=category,
            )
        )
    return rows


def _parse_csv_metadata(text: str, csv_url: str) -> list[dict[str, str]]:
    """Extract series metadata from an RBA CSV file's header rows.

    RBA CSVs carry a fixed-shape metadata block before the data rows (Title /
    Description / Frequency / Type / Units / Source / Publication date / Series ID).
    One catalog row per ``Series ID`` column; Description is the highest-signal field
    for retrieval since it spells out the measure in human-readable English.
    """
    lines = text.strip().split("\n")
    content_lines = [ln for ln in lines[1:] if ln.strip()]
    if len(content_lines) < 8:
        return []

    try:
        reader = pd.read_csv(io.StringIO("\n".join(content_lines)), header=None, dtype=str, nrows=10)
    except Exception:
        return []

    if reader.empty or len(reader) < 8:
        return []

    title_row = description_row = frequency_row = units_row = series_id_row_idx = None
    for i in range(min(10, len(reader))):
        first_val = str(reader.iloc[i, 0]).strip() if pd.notna(reader.iloc[i, 0]) else ""
        if first_val == "Title":
            title_row = i
        elif first_val == "Description":
            description_row = i
        elif first_val == "Frequency":
            frequency_row = i
        elif first_val == "Units":
            units_row = i
        elif first_val == "Series ID":
            series_id_row_idx = i

    if series_id_row_idx is None or title_row is None:
        return []

    csv_filename = csv_url.split("/")[-1].replace(".csv", "")
    category = _CATEGORY_PREFIXES.get(csv_filename[0].lower() if csv_filename else "", "")

    def _cell(row_idx: int | None, col: int) -> str:
        if row_idx is None:
            return ""
        val = reader.iloc[row_idx, col]
        if not pd.notna(val):
            return ""
        return str(val).strip()

    rows: list[dict[str, str]] = []
    for col in reader.columns[1:]:
        sid = _cell(series_id_row_idx, col)
        if not sid or sid == "nan":
            continue
        title = _cell(title_row, col) or sid
        description = _cell(description_row, col)
        frequency = _cell(frequency_row, col)
        unit = _cell(units_row, col)

        rows.append(
            {
                "code": f"{csv_filename}#{sid}",
                "table_id": csv_filename,
                "series_id": sid,
                "title": title,
                "description": description,
                "category": category,
                "frequency": frequency,
                "unit": unit,
                "source": "rba_csv",
            }
        )

    return rows
