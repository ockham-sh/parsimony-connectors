"""Reserve Bank of Australia (RBA): fetch + catalog enumeration.

Data: https://www.rba.gov.au/statistics/tables/. No authentication required
(keyless public statistics site — no ``secrets=``/``bind()``/``UnauthorizedError``;
``load()`` binds only the catalog URL for search).

Transport — the Akamai/``curl_cffi`` special case
-------------------------------------------------
``rba.gov.au`` is fronted by Akamai bot-mitigation that **TLS-fingerprint-blocks
stock python-httpx** — every request through the canonical
``make_http_client``/``fetch_json`` path returns HTTP 403. The canonical
transport therefore structurally *cannot reach this host*, which is exactly the
§6 sanctioned "raw transport + custom error mapper" exception. RBA fetches go
through **curl_cffi** (``Session(...).get(url, impersonate="chrome")``),
which presents a real Chrome TLS handshake and gets HTTP 200. curl_cffi is a
HARD dependency (declared in ``pyproject.toml``): without it the connector is
non-functional.

Because curl_cffi is not httpx, the kernel ``map_http_error`` / ``map_timeout_error``
helpers don't apply. :func:`_curl_get` is the hand-written mapper required by §6:
it inspects ``response.status_code`` and maps to the typed-error taxonomy
(429 → :class:`RateLimitError`, 402 → :class:`PaymentRequiredError`,
401/403 → :class:`UnauthorizedError`, other 4xx/5xx → :class:`ProviderError`),
and converts curl_cffi timeout/connection failures to
``ProviderError(status_code=408)`` — mirroring ``map_timeout_error``.

Discovery runs three passes across RBA's distinct publication formats:

1. **CSV index** (``/statistics/tables/``): ~216 CSVs, ~3,957 active series.
   Parses the metadata header rows (Title / Description / Units /
   Frequency / Series ID). This is the bulk of the catalog.
2. **XLSX-exclusive sub-sheets** (``/statistics/tables/xls/``): workbooks
   that mirror the CSVs *except* for a handful of sheets never republished
   as CSV. Right now only ``a03.xlsx`` — ``Bond Purchase Program`` (+7
   series: face value, coupon, cut-off rate, WAR, etc.) qualifies; the
   allow-list is explicit so we don't double-count.
3. **Legacy xls-hist binaries** (``/statistics/historical-data.html``):
   ~37 ``.xls`` files with discontinued series that left the live CSVs
   — b3 repo rates, c9 cheque/card historical, e4-e7 household finance,
   f16 retail interest, etc. Parsed with xlrd since they predate xlsx.
   ~186 series that are otherwise uncatalogable.

The fetch connector accepts a CSV filename (e.g. ``f1-data``) as the
``table_id`` and resolves it against the live tables page. This avoids
hard-coding URL patterns that break when the RBA renames files.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from datetime import datetime
from typing import TYPE_CHECKING, Annotated, Any
from xml.etree import ElementTree as ET

import openpyxl
import pandas as pd
import xlrd
from curl_cffi.requests import Session
from curl_cffi.requests import exceptions as curl_exc
from parsimony import Namespace
from parsimony.connector import Connectors, connector, enumerator
from parsimony.errors import (
    EmptyDataError,
    InvalidParameterError,
    PaymentRequiredError,
    ProviderError,
    RateLimitError,
    UnauthorizedError,
)
from parsimony.result import (
    Column,
    ColumnRole,
    OutputConfig,
)
from parsimony.transport import parse_retry_after

if TYPE_CHECKING:  # pragma: no cover - typing only
    from curl_cffi.requests import Response

logger = logging.getLogger(__name__)

_BASE_URL = "https://www.rba.gov.au"
_TABLES_URL = f"{_BASE_URL}/statistics/tables/"
_HISTORICAL_URL = f"{_BASE_URL}/statistics/historical-data.html"
_CSV_LINK_PATTERN = re.compile(r'href="(/statistics/tables/csv/([^"]+)\.csv)"')
_XLSX_LINK_PATTERN = re.compile(r'href="/statistics/tables/xls/([^"]+)\.xlsx"')
_XLS_HIST_LINK_PATTERN = re.compile(r'href="/statistics/tables/xls-hist/([^"]+)\.xls"')

#: Per-request timeout (seconds) for the curl_cffi GETs. The curl_cffi GET
#: impersonates a recent Chrome TLS fingerprint (``impersonate="chrome"``) so
#: Akamai lets the request through; older fingerprints have started to 403.
_TIMEOUT = 60.0

# Allow-list of ``xls/<stem>.xlsx`` sub-sheets whose series are NOT
# republished in any CSV. Key = xlsx filename stem; value = tuple of
# sheet names to scan. Sheets NOT in this allow-list are skipped to avoid
# double-counting series already captured from CSV. Extend as new
# XLSX-only sheets are identified.
_XLSX_EXCLUSIVE_SHEETS: dict[str, tuple[str, ...]] = {
    # a03.xlsx's Bond Purchase Program sheet carries 7 series (bond issuer,
    # coupon, maturity, face value, WAR, cut-off rate, value date) that the
    # a3-* CSVs do not publish.
    "a03": ("Bond Purchase Program",),
}

_CATEGORY_PREFIXES = {
    "a": "Reserve Bank",
    "b": "Banking and Finance",
    "c": "Credit and Charge Cards",
    "d": "Monetary Aggregates",
    "e": "Household and Business Finance",
    "f": "Interest Rates and Yields",
    "g": "Exchange Rates",
    "h": "Economic Activity",
    "i": "Balance of Payments",
}


# ---------------------------------------------------------------------------
# Output configs
# ---------------------------------------------------------------------------

RBA_ENUMERATE_OUTPUT = OutputConfig(
    columns=[
        # Compound code ``{table_id}#{series_id}`` so every series gets a unique
        # catalog entry. RBA reuses some series IDs across closely-related
        # tables (e.g. ``b13.1.2-africa-and-middle-east`` vs
        # ``b13.2.1-africa-and-middle-east`` share ~225 ids each); a bare
        # ``series_id`` KEY would silently dedup ~5% of entries. Mirrors
        # Treasury's ``{endpoint}#{field}`` precedent. Agents split on ``#``
        # to recover the fetchable ``table_id`` and the row's ``series_key``.
        Column(name="code", role=ColumnRole.KEY, namespace="rba"),
        Column(name="title", role=ColumnRole.TITLE),
        # ``description`` is the CSV header's own per-series descriptive text
        # — the most useful semantic signal for retrieval.
        Column(name="description", role=ColumnRole.METADATA),
        # ``source`` tells the agent which fetch connector to call —
        # currently a single source (``"rba_csv"`` → :func:`rba_fetch`),
        # declared explicitly so dispatch stays consistent if/when more
        # RBA data shapes are added.
        Column(name="source", role=ColumnRole.METADATA),
        Column(name="table_id", role=ColumnRole.METADATA),
        Column(name="series_id", role=ColumnRole.METADATA),
        Column(name="category", role=ColumnRole.METADATA),
        Column(name="frequency", role=ColumnRole.METADATA),
        Column(name="unit", role=ColumnRole.METADATA),
    ]
)

RBA_FETCH_OUTPUT = OutputConfig(
    columns=[
        Column(name="table_id", role=ColumnRole.KEY, namespace="rba"),
        Column(name="title", role=ColumnRole.TITLE),
        Column(name="date", dtype="datetime", role=ColumnRole.DATA),
        Column(name="value", dtype="numeric", role=ColumnRole.DATA),
        Column(name="series_key", role=ColumnRole.DATA),
    ]
)


# ---------------------------------------------------------------------------
# HTTP — curl_cffi raw transport + a hand-written error mapper (§6 exception)
# ---------------------------------------------------------------------------
#
# Stock httpx 403s against RBA's Akamai TLS-fingerprint check, so the canonical
# make_http_client/fetch_json path can't be used. curl_cffi impersonates Chrome
# and reaches the origin; because it is NOT httpx, we map its failures to the
# typed-error taxonomy by hand here — the §6 "raw + custom mapper" carve-out for
# a host the canonical transport structurally cannot reach.


def _map_curl_status(status_code: int, *, response: Response, op_name: str) -> None:
    """Map a non-2xx curl_cffi status to the typed-error taxonomy (NoReturn-style).

    Mirrors the kernel ``map_http_error`` status table. RBA is keyless, so
    401/403 are vanishingly unlikely on the data paths (Akamai blocks *before*
    auth and we impersonate a browser), but we map them faithfully for
    completeness rather than letting them fall through to a generic
    ProviderError.
    """
    if status_code == 429:
        raise RateLimitError("rba", retry_after=_retry_after_seconds(response))
    if status_code == 402:
        raise PaymentRequiredError("rba")
    if status_code in (401, 403):
        raise UnauthorizedError("rba")
    raise ProviderError("rba", status_code=status_code)


def _retry_after_seconds(response: Response, *, default: float = 60.0) -> float:
    """Parse a ``Retry-After`` duration (seconds) from a curl_cffi response.

    The kernel ``parse_retry_after`` is typed for ``httpx.Response`` only; the
    curl_cffi response has a duck-compatible ``headers.get`` but a different
    static type, so we reuse the kernel parser via a tiny shim object that
    presents just the ``.headers`` attribute it reads.
    """
    raw = response.headers.get("Retry-After", "")
    if not str(raw).strip():
        return default

    class _Shim:
        headers = response.headers

    return parse_retry_after(_Shim(), default=default)  # type: ignore[arg-type]


def _curl_get(session: Session, url: str, *, op_name: str, binary: bool = False) -> str | bytes:
    """GET *url* via curl_cffi (Chrome impersonation) → text or bytes.

    The raw §6 transport for an Akamai-blocked host: issue the GET, inspect
    ``response.status_code`` directly, and map any non-2xx through
    :func:`_map_curl_status`. curl_cffi timeout / connection failures map to
    ``ProviderError(status_code=408)`` (the ``map_timeout_error`` convention).
    The body is parsed separately by the caller.
    """
    try:
        response = session.get(url, impersonate="chrome", timeout=_TIMEOUT)
    except curl_exc.Timeout as exc:
        raise ProviderError("rba", status_code=408) from exc
    except curl_exc.RequestException as exc:
        # ConnectionError / DNSError / SSLError / ImpersonateError / etc. —
        # treat any transport-level failure as a transient provider error so
        # the agent can pick another connector (timeout bucket, 408).
        raise ProviderError("rba", status_code=408) from exc

    if response.status_code >= 400:
        _map_curl_status(response.status_code, response=response, op_name=op_name)

    if binary:
        content = response.content
        return content if isinstance(content, bytes) else bytes(content)
    return str(response.text)


def _make_session() -> Session:
    """Build a curl_cffi session. One per ``rba_fetch`` call; reused (pooled)
    across the enumerator's fan-out."""
    return Session()


# ---------------------------------------------------------------------------
# URL resolution
# ---------------------------------------------------------------------------


def _discover_csv_links(session: Session) -> list[str]:
    """Scrape the tables index and return the CSV link paths it advertises.

    The **primary bounding seam** for the enumerator's live test: monkeypatch
    this module-global to return a 1–2 link slice and the CSV fan-out fires a
    handful of requests instead of the full ~216-CSV crawl. The enumerator
    reads it as a module global at call time, so the monkeypatch takes.
    """
    html = _curl_get(session, _TABLES_URL, op_name="tables_index")
    assert isinstance(html, str)
    return [m[0] for m in _CSV_LINK_PATTERN.findall(html)]


def _discover_xlsx_stems(session: Session) -> set[str]:
    """Scrape the tables index and return the XLSX workbook stems present.

    A **bounding seam** (live tests monkeypatch this to ``set()`` so the
    XLSX-exclusive pass fires zero requests). Only stems in
    :data:`_XLSX_EXCLUSIVE_SHEETS` are actually fetched downstream.
    """
    html = _curl_get(session, _TABLES_URL, op_name="tables_index")
    assert isinstance(html, str)
    return set(_XLSX_LINK_PATTERN.findall(html))


def _discover_xls_hist_stems(session: Session) -> list[str]:
    """Scrape the historical-data index and return the xls-hist workbook stems.

    A **bounding seam** (live tests monkeypatch this to ``[]`` so the xls-hist
    pass — which otherwise fetches ~37 legacy ``.xls`` binaries — fires zero
    requests). Read as a module global at call time so the monkeypatch takes.
    """
    try:
        hist_html = _curl_get(session, _HISTORICAL_URL, op_name="historical_index")
        assert isinstance(hist_html, str)
    except Exception:
        return []
    return sorted(set(_XLS_HIST_LINK_PATTERN.findall(hist_html)))


def _resolve_csv_url(session: Session, table_id: str) -> str:
    """Scrape the RBA tables page and resolve *table_id* to a full CSV URL.

    Matches the table_id against known CSV filenames on the page.
    """
    html = _curl_get(session, _TABLES_URL, op_name="tables_index")
    assert isinstance(html, str)
    matches = _CSV_LINK_PATTERN.findall(html)

    # Build lookup: filename stem (lowercase) → full path
    stem_to_path: dict[str, str] = {}
    for path, stem in matches:
        stem_to_path[stem.lower()] = path

    tid = table_id.lower()

    # Exact match
    if tid in stem_to_path:
        return f"{_BASE_URL}{stem_to_path[tid]}"

    # Fuzzy: caller might use "f1" when actual stem is "f1-data"
    for stem, path in stem_to_path.items():
        if stem.startswith(tid + "-") or stem == tid:
            return f"{_BASE_URL}{path}"

    available = sorted(stem_to_path.keys())[:20]
    raise InvalidParameterError(
        "rba",
        f"RBA table '{table_id}' not found. Available tables include: {', '.join(available)}...",
    )


# ---------------------------------------------------------------------------
# CSV parsing helpers
# ---------------------------------------------------------------------------


def _normalize_date(s: str) -> str:
    for fmt in ("%d-%b-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _parse_rba_csv(text: str, table_id: str) -> pd.DataFrame:
    """Parse RBA CSV: skip metadata header, melt to long format."""
    lines = text.strip().split("\n")

    header_idx = 0
    series_id_row: dict[str, str] = {}

    for i, line in enumerate(lines):
        lower = line.strip().lower()
        if lower.startswith("series id"):
            reader = csv.reader(io.StringIO(line))
            parts = list(reader)[0]
            for j, part in enumerate(parts):
                series_id_row[str(j)] = part.strip()
        if lower.startswith("title"):
            header_idx = i

    if header_idx == 0:
        header_idx = min(10, len(lines) - 2)

    data_text = "\n".join(lines[header_idx:])
    rows_list = list(csv.reader(io.StringIO(data_text)))

    if len(rows_list) < 2:
        return pd.DataFrame(columns=["table_id", "title", "date", "value", "series_key"])

    header = rows_list[0]
    data_start = 1
    for i, row in enumerate(rows_list):
        if row and row[0].strip().lower().startswith("series id"):
            data_start = i + 1
            break

    all_rows: list[dict[str, Any]] = []

    for row in rows_list[data_start:]:
        if not row or not row[0].strip():
            continue
        date = _normalize_date(row[0].strip())
        for col_idx in range(1, min(len(header), len(row))):
            col_name = header[col_idx].strip()
            if not col_name:
                continue
            val_str = row[col_idx].strip()
            try:
                value = float(val_str) if val_str else None
            except (ValueError, TypeError):
                value = None
            all_rows.append(
                {
                    "table_id": table_id,
                    "title": col_name,
                    "date": date,
                    "value": value,
                    "series_key": series_id_row.get(str(col_idx), col_name),
                }
            )

    return (
        pd.DataFrame(all_rows)
        if all_rows
        else pd.DataFrame(columns=["table_id", "title", "date", "value", "series_key"])
    )


def _metadata_from_header_rows(
    sheet_rows: list[list[Any]],
    *,
    table_id: str,
    sheet_name: str,
    source: str,
    category: str,
) -> list[dict[str, str]]:
    """Extract series metadata from RBA-shaped header rows.

    Shared post-parsing step for both XLSX (openpyxl) and XLS-hist (xlrd)
    workbooks. ``sheet_rows`` is the first ~15 rows of a sheet, each a
    list of cell values with column 0 carrying the label (``Title``,
    ``Description``, ..., ``Series ID`` / ``Mnemonic``). XLSX workbooks
    use ``Series ID``; the legacy ``xls-hist/zcr-analytical-series-hist``
    uses ``Mnemonic`` instead — handled transparently here. One catalog
    row per non-empty id column.
    """
    title_row: list[Any] | None = None
    description_row: list[Any] | None = None
    frequency_row: list[Any] | None = None
    units_row: list[Any] | None = None
    id_row: list[Any] | None = None
    for row in sheet_rows[:20]:
        if not row:
            continue
        head = str(row[0]).strip() if row[0] is not None else ""
        if head == "Title":
            title_row = row
        elif head == "Description":
            description_row = row
        elif head == "Frequency":
            frequency_row = row
        elif head == "Units":
            units_row = row
        elif head in ("Series ID", "Mnemonic"):
            id_row = row

    if not id_row:
        return []

    def _cell(row: list[Any] | None, col: int) -> str:
        if row is None or col >= len(row):
            return ""
        val = row[col]
        if val is None:
            return ""
        s = str(val).strip()
        return "" if s.lower() == "nan" else s

    rows: list[dict[str, str]] = []
    for col in range(1, len(id_row)):
        sid = _cell(id_row, col)
        if not sid:
            continue
        title = _cell(title_row, col) or sid
        description = _cell(description_row, col)
        frequency = _cell(frequency_row, col)
        unit = _cell(units_row, col)
        # When a workbook has multiple data sheets, disambiguate by
        # folding the sheet name into the table_id so rows stay unique
        # even if the same series id is duplicated across sheets.
        effective_table_id = table_id if not sheet_name else f"{table_id}/{sheet_name}"
        rows.append(
            {
                "code": f"{effective_table_id}#{sid}",
                "table_id": effective_table_id,
                "series_id": sid,
                "title": title,
                "description": description,
                "category": category,
                "frequency": frequency,
                "unit": unit,
                "source": source,
            }
        )
    return rows


def _xlsx_sheet_rows(wb: openpyxl.Workbook, sheet_name: str) -> list[list[Any]]:
    """First 20 rows of an XLSX sheet as plain lists of cell values."""
    ws = wb[sheet_name]
    rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 20:
            break
        rows.append(list(row))
    return rows


def _xls_sheet_rows(sheet: xlrd.sheet.Sheet) -> list[list[Any]]:
    """First 20 rows of a legacy .xls sheet as plain lists."""
    rows: list[list[Any]] = []
    for r in range(min(20, sheet.nrows)):
        rows.append(list(sheet.row_values(r)))
    return rows


def _parse_xlsx_exclusive(
    data: bytes,
    table_id: str,
    allowed_sheets: tuple[str, ...],
) -> list[dict[str, str]]:
    """Parse metadata rows from an XLSX workbook's allow-listed sheets.

    Skips non-matching sheets so we don't re-emit series already captured
    from the CSV index. Used for :data:`_XLSX_EXCLUSIVE_SHEETS`.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return []
    category = _CATEGORY_PREFIXES.get(table_id[0].lower() if table_id else "", "")
    rows: list[dict[str, str]] = []
    for sheet_name in allowed_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        sheet_rows = _xlsx_sheet_rows(wb, sheet_name)
        rows.extend(
            _metadata_from_header_rows(
                sheet_rows,
                table_id=table_id,
                sheet_name=sheet_name,
                source="rba_xlsx",
                category=category,
            )
        )
    wb.close()
    return rows


def _parse_xls_hist(data: bytes, table_id: str) -> list[dict[str, str]]:
    """Parse a legacy ``.xls`` workbook from ``/statistics/tables/xls-hist/``.

    Walks every data sheet (skipping ``Notes`` / ``Series breaks``) and
    emits one row per series id column. Series are suffixed with the
    sheet name to keep keys unique when a workbook has multiple data
    sheets (e.g. ``a03hist-2003-2008.xls`` has one sheet per security).
    """
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception:
        return []
    # Category prefix uses the base table letter (before any ``hist``
    # suffix) so ``b03hist`` maps to the same "Banking and Finance" as
    # ``b3-data`` from CSV.
    base = table_id.lstrip("_").lower()
    for digit_marker in ("hist", "historical", "0", "1", "2"):
        idx = base.find(digit_marker)
        if idx > 0:
            base = base[:idx]
            break
    category = _CATEGORY_PREFIXES.get(table_id[0].lower() if table_id else "", "")

    rows: list[dict[str, str]] = []
    for sheet_name in wb.sheet_names():
        low = sheet_name.lower()
        if "note" in low or "series breaks" in low:
            continue
        sheet = wb.sheet_by_name(sheet_name)
        sheet_rows = _xls_sheet_rows(sheet)
        # Heuristic: only attach sheet_name to table_id when the workbook
        # has >1 data sheet. Single-sheet workbooks keep a cleaner code.
        attach_sheet = (
            sum(1 for sn in wb.sheet_names() if "note" not in sn.lower() and "series breaks" not in sn.lower()) > 1
        )
        rows.extend(
            _metadata_from_header_rows(
                sheet_rows,
                table_id=table_id,
                sheet_name=sheet_name if attach_sheet else "",
                source="rba_xlsx_hist",
                category=category,
            )
        )
    return rows


def _parse_xlsx_raw_sheet(data: bytes, sheet_index: int = 1) -> list[list[str]]:
    """Fallback XLSX parser that bypasses openpyxl's stylesheet validation.

    A handful of RBA XLSX workbooks (``b11-1-hist``, ``c07hist``,
    ``d05hist``, etc.) ship with a ``[trash]/`` folder and a stylesheet
    openpyxl refuses to open. Those files are valid zips otherwise, so
    this reads ``xl/sharedStrings.xml`` + ``xl/worksheets/sheet{N}.xml``
    directly and returns the first 20 rows as lists of strings. Used
    only when openpyxl fails, since the layout heuristic is fragile.
    """
    import zipfile

    try:
        z = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        return []
    try:
        ss_xml = z.read("xl/sharedStrings.xml").decode("utf-8")
        sheet_xml = z.read(f"xl/worksheets/sheet{sheet_index}.xml").decode("utf-8")
    except KeyError:
        return []
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ss_root = ET.fromstring(ss_xml)
    strings: list[str] = []
    for si in ss_root.findall(f"{ns}si"):
        t_elem = si.find(f"{ns}t")
        strings.append(t_elem.text if (t_elem is not None and t_elem.text) else "")

    sh = ET.fromstring(sheet_xml)
    data_elt = sh.find(f"{ns}sheetData")
    if data_elt is None:
        return []
    rows: list[list[str]] = []
    for i, row in enumerate(data_elt.findall(f"{ns}row")):
        if i >= 20:
            break
        cells: list[str] = []
        for c in row.findall(f"{ns}c"):
            cell_type = c.attrib.get("t")
            v = c.find(f"{ns}v")
            if v is None:
                cells.append("")
                continue
            if cell_type == "s":
                try:
                    cells.append(strings[int(v.text or "0")])
                except (ValueError, IndexError):
                    cells.append("")
            else:
                cells.append(v.text or "")
        rows.append(cells)
    return rows


def _parse_csv_metadata(text: str, csv_url: str) -> list[dict[str, str]]:
    """Extract series metadata from an RBA CSV file's header rows.

    RBA CSVs carry a fixed-shape metadata block before the data rows
    (Title / Description / Frequency / Type / Units / Source / Publication
    date / Series ID). One catalog row per ``Series ID`` column; the
    enumerator captures Title, Description (rich semantic text), Frequency,
    and Units, and derives ``table_id`` from the URL. Description is the
    highest-signal field for retrieval since it spells out the measure
    in human-readable English.
    """
    lines = text.strip().split("\n")
    content_lines = [ln for ln in lines[1:] if ln.strip()]
    if len(content_lines) < 8:
        return []

    try:
        reader = pd.read_csv(io.StringIO("\n".join(content_lines)), header=None, dtype=str, nrows=10)
    except Exception:
        return []

    if reader.empty or len(reader) < 8:
        return []

    title_row = description_row = frequency_row = units_row = series_id_row_idx = None
    for i in range(min(10, len(reader))):
        first_val = str(reader.iloc[i, 0]).strip() if pd.notna(reader.iloc[i, 0]) else ""
        if first_val == "Title":
            title_row = i
        elif first_val == "Description":
            description_row = i
        elif first_val == "Frequency":
            frequency_row = i
        elif first_val == "Units":
            units_row = i
        elif first_val == "Series ID":
            series_id_row_idx = i

    if series_id_row_idx is None or title_row is None:
        return []

    csv_filename = csv_url.split("/")[-1].replace(".csv", "")
    category = _CATEGORY_PREFIXES.get(csv_filename[0].lower() if csv_filename else "", "")

    def _cell(row_idx: int | None, col: int) -> str:
        if row_idx is None:
            return ""
        val = reader.iloc[row_idx, col]
        if not pd.notna(val):
            return ""
        return str(val).strip()

    rows: list[dict[str, str]] = []
    for col in reader.columns[1:]:
        sid = _cell(series_id_row_idx, col)
        if not sid or sid == "nan":
            continue
        title = _cell(title_row, col) or sid
        description = _cell(description_row, col)
        frequency = _cell(frequency_row, col)
        unit = _cell(units_row, col)

        rows.append(
            {
                # ``code`` is the compound catalog key — see RBA_ENUMERATE_OUTPUT.
                "code": f"{csv_filename}#{sid}",
                "table_id": csv_filename,
                "series_id": sid,
                "title": title,
                "description": description,
                "category": category,
                "frequency": frequency,
                "unit": unit,
                "source": "rba_csv",
            }
        )

    return rows


# ---------------------------------------------------------------------------
# Connectors
# ---------------------------------------------------------------------------


@connector(output=RBA_FETCH_OUTPUT, tags=["macro", "au"])
def rba_fetch(table_id: Annotated[str, Namespace("rba")]) -> pd.DataFrame:
    """Fetch RBA statistical table data by table ID.

    Resolves the table_id against the live RBA tables page to find the
    correct CSV URL, then downloads and parses the data. The RBA site is
    Akamai-protected, so the fetch uses curl_cffi (browser impersonation);
    plain httpx is blocked at the TLS layer.
    """
    table_id = table_id.strip().lower()
    if table_id.endswith(".csv"):
        table_id = table_id[:-4]
    if not table_id:
        raise InvalidParameterError("rba", "table_id must be non-empty")

    with _make_session() as session:
        url = _resolve_csv_url(session, table_id)
        text = _curl_get(session, url, op_name="csv")
    assert isinstance(text, str)

    try:
        df = _parse_rba_csv(text, table_id)
    except Exception as exc:  # noqa: BLE001 — any structural parse failure → ParseError (§5.8)
        from parsimony.errors import ParseError

        raise ParseError("rba", f"could not parse RBA CSV for table {table_id!r}: {exc}") from exc

    if df.empty:
        raise EmptyDataError(
            "rba",
            message=f"No data returned for table: {table_id}",
            query_params={"table_id": table_id},
        )

    return df


_ENUMERATE_COLUMNS: tuple[str, ...] = (
    "code",
    "title",
    "description",
    "source",
    "table_id",
    "series_id",
    "category",
    "frequency",
    "unit",
)


@enumerator(output=RBA_ENUMERATE_OUTPUT, tags=["macro", "au"])
def enumerate_rba() -> pd.DataFrame:
    """Discover RBA series from CSV index, XLSX sheets, and xls-hist files.

    Compound catalog keys use ``table_id#series_id`` so duplicate series ids
    across tables remain addressable without losing descriptions. The serial
    crawl reuses one curl_cffi session across all ~250 requests
    (Akamai-impersonated pooling); per-fetch failures are skipped, not fatal.
    """
    all_rows: list[dict[str, str]] = []
    seen: set[str] = set()

    with _make_session() as session:
        # Step 1: discover the three link sets via module-global seams. Each is
        # read at call time so a live test can monkeypatch any of them to bound
        # the fan-out (the CSV seam is the primary one; the XLSX/xls-hist seams
        # bound the secondary passes).
        csv_links = _discover_csv_links(session)
        try:
            xlsx_stems = _discover_xlsx_stems(session)
        except Exception:
            xlsx_stems = set()

        if not csv_links and not xlsx_stems:
            return pd.DataFrame(columns=list(_ENUMERATE_COLUMNS))

        def _fetch_csv(link: str) -> list[dict[str, str]]:
            url = f"{_BASE_URL}{link}"
            try:
                text = _curl_get(session, url, op_name="csv_metadata")
                assert isinstance(text, str)
                return _parse_csv_metadata(text, url)
            except Exception:
                return []

        # Step 2: CSV pass.
        for rows in [_fetch_csv(link) for link in csv_links]:
            for row in rows:
                code = row["code"]
                if code not in seen:
                    seen.add(code)
                    all_rows.append(row)

        # Step 3: XLSX-exclusive sub-sheet pass. Only fetches workbooks in
        # the allow-list, not the full ~86-file XLSX set — the rest duplicate
        # CSV content.
        def _fetch_xlsx(stem: str, sheets: tuple[str, ...]) -> list[dict[str, str]]:
            url = f"{_BASE_URL}/statistics/tables/xls/{stem}.xlsx"
            try:
                data = _curl_get(session, url, op_name="xlsx", binary=True)
                assert isinstance(data, bytes)
                return _parse_xlsx_exclusive(data, stem, sheets)
            except Exception:
                return []

        xlsx_targets = [(stem, sheets) for stem, sheets in _XLSX_EXCLUSIVE_SHEETS.items() if stem in xlsx_stems]
        for rows in [_fetch_xlsx(stem, sheets) for stem, sheets in xlsx_targets]:
            for row in rows:
                code = row["code"]
                if code not in seen:
                    seen.add(code)
                    all_rows.append(row)

        # Step 4: xls-hist legacy binary pass. The historical-data.html index
        # lists ~37 .xls workbooks covering discontinued series.
        xls_hist_stems = _discover_xls_hist_stems(session)

        def _fetch_xls_hist(stem: str) -> list[dict[str, str]]:
            url = f"{_BASE_URL}/statistics/tables/xls-hist/{stem}.xls"
            try:
                data = _curl_get(session, url, op_name="xls_hist", binary=True)
                assert isinstance(data, bytes)
                return _parse_xls_hist(data, stem)
            except Exception:
                return []

        # Skip obvious "period range" workbooks (``1983-1986.xls``, etc.)
        # whose sheets lack Series ID rows — they return empty and cost a
        # fetch. Detection: first character isn't a letter.
        hist_targets = [stem for stem in xls_hist_stems if stem and stem[0].isalpha()]
        for rows in [_fetch_xls_hist(stem) for stem in hist_targets]:
            for row in rows:
                code = row["code"]
                if code not in seen:
                    seen.add(code)
                    all_rows.append(row)

    df = (
        pd.DataFrame(all_rows, columns=list(_ENUMERATE_COLUMNS))
        if all_rows
        else pd.DataFrame(columns=list(_ENUMERATE_COLUMNS))
    )
    return df


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------

from parsimony_rba.search import (  # noqa: E402, F401  (after public decorators; re-exported)
    PARSIMONY_RBA_CATALOG_URL_ENV,
    RBA_SEARCH_OUTPUT,
    rba_search,
)

CONNECTORS = Connectors([rba_fetch, enumerate_rba, rba_search])


def load(*, catalog_url: str | None = None) -> Connectors:
    """Return :data:`CONNECTORS` with an optional catalog-search URL bound.

    RBA is keyless, so there is no API key to bind — only the catalog
    snapshot URL for ``rba_search`` (overrides the published default /
    ``PARSIMONY_RBA_CATALOG_URL`` env var when supplied).
    """
    if catalog_url is None:
        return CONNECTORS
    return CONNECTORS.bind(catalog_url=catalog_url)
